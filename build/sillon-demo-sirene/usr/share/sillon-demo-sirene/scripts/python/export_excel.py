"""SILLON - Script Python d'exemple : export Excel multi-feuilles avec graphique natif (sillon-demo-sirene).

Classeur openpyxl à trois feuilles (résumé, top secteurs NAF, par
département) construit à partir de requêtes déjà agrégées côté PostgreSQL
(jamais un SELECT * sur les ~43,9 millions de lignes) - openpyxl.chart
permet d'insérer un graphique natif directement lisible/modifiable dans
Excel ou LibreOffice, pas seulement une image figée.
"""
import os

import psycopg2
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

dsn = os.environ["SILLON_DSN"]
resultats = os.environ["SILLON_RESULTATS"]

connexion = psycopg2.connect(dsn)
curseur = connexion.cursor()

curseur.execute("SELECT COUNT(*) FROM sirene_etablissements")
total = curseur.fetchone()[0]
curseur.execute("SELECT COUNT(*) FROM sirene_etablissements WHERE etat_administratif = 'A'")
actifs = curseur.fetchone()[0]
curseur.execute("SELECT COUNT(*) FROM sirene_etablissements WHERE etablissement_siege")
sieges = curseur.fetchone()[0]

curseur.execute("""
    SELECT LEFT(activite_principale, 2) AS division_naf, COUNT(*) AS nb
    FROM sirene_etablissements
    WHERE etat_administratif = 'A' AND activite_principale IS NOT NULL AND activite_principale != ''
    GROUP BY division_naf ORDER BY nb DESC LIMIT 20
""")
top_secteurs = curseur.fetchall()

curseur.execute("""
    SELECT dep_code, COUNT(*) AS nb FROM sirene_etablissements
    WHERE etat_administratif = 'A' AND dep_code != '' GROUP BY dep_code ORDER BY dep_code
""")
par_departement = curseur.fetchall()
connexion.close()

classeur = Workbook()

feuille_resume = classeur.active
feuille_resume.title = "Résumé"
feuille_resume.append(["Indicateur", "Valeur"])
for cellule in feuille_resume[1]:
    cellule.font = Font(bold=True)
feuille_resume.append(["Établissements (total)", total])
feuille_resume.append(["Établissements actifs", actifs])
feuille_resume.append(["Sièges actifs", sieges])

feuille_secteurs = classeur.create_sheet("Top secteurs NAF")
feuille_secteurs.append(["Division NAF", "Établissements actifs"])
for cellule in feuille_secteurs[1]:
    cellule.font = Font(bold=True)
for division, nb in top_secteurs:
    feuille_secteurs.append([division, nb])

graphique = BarChart()
graphique.title = "20 divisions NAF les plus représentées (actifs)"
graphique.y_axis.title = "Établissements actifs"
donnees = Reference(feuille_secteurs, min_col=2, min_row=1, max_row=1 + len(top_secteurs))
categories = Reference(feuille_secteurs, min_col=1, min_row=2, max_row=1 + len(top_secteurs))
graphique.add_data(donnees, titles_from_data=True)
graphique.set_categories(categories)
feuille_secteurs.add_chart(graphique, "E2")

feuille_dep = classeur.create_sheet("Par département")
feuille_dep.append(["Département", "Établissements actifs"])
for cellule in feuille_dep[1]:
    cellule.font = Font(bold=True)
for dep_code, nb in par_departement:
    feuille_dep.append([dep_code, nb])

classeur.save(os.path.join(resultats, "sirene_synthese.xlsx"))

print(f"Classeur produit : {total:,} établissements, {actifs:,} actifs, {sieges:,} sièges.".replace(",", " "))
